#!/usr/bin/env python3
"""Export the Bert Airtable base schema to a review workbook.

The Airtable MCP tools split the schema across two calls, and neither is
sufficient alone:

  * list_tables_for_base  -> field NAMES and DESCRIPTIONS (no select choices)
  * get_table_schema      -> select CHOICES and type config (no names at all)

So this takes both payloads as saved JSON and merges them. Refresh a snapshot by
re-running those two MCP calls and saving their raw JSON, then:

    python3 scripts/bert_schema_export.py \\
        --names names.json --config config.json \\
        --out analysis/output/bert-airtable-schema-YYYY-MM-DD.xlsx

Sheets produced:
  Overview       one row per table
  <table name>   one row per field: name, id, type, description, choices
  Select Options long format, one row per select choice — the sheet to review
                 data-entry vocabularies in
  Review Flags   options that look like import artifacts, and undocumented fields

Read-only. Writes only the output workbook (and its JSON snapshot if asked).
"""

import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RIDER_BLUE = "1E6FFF"  # Titan brand colour, per titan-agents CLAUDE.md
FONT = "Arial"
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")


def load_merged(names_path: Path, config_path: Path) -> list[dict]:
    """Merge the two MCP payloads into one table list, keyed by field id."""
    names = json.loads(names_path.read_text())["tables"]
    config = {t["tableId"]: t for t in json.loads(config_path.read_text())["tables"]}

    merged = []
    for tbl in names:
        cfg_fields = {f["id"]: f for f in config.get(tbl["id"], {}).get("fields", [])}
        fields = []
        for i, f in enumerate(tbl["fields"], start=1):
            cfg = cfg_fields.get(f["id"], {}).get("config", {}) or {}
            fields.append({
                "order": i,
                "id": f["id"],
                "name": f["name"],
                "type": f["type"],
                "description": (f.get("description") or "").strip(),
                "choices": [c.get("name", "") for c in cfg.get("choices", [])],
                "choice_ids": [c.get("id", "") for c in cfg.get("choices", [])],
                "choice_colors": [c.get("color", "") for c in cfg.get("choices", [])],
                "config": {k: v for k, v in cfg.items() if k != "choices"},
            })
        merged.append({
            "id": tbl["id"],
            "name": tbl["name"],
            "description": (tbl.get("description") or "").strip(),
            "primaryFieldId": tbl.get("primaryFieldId", ""),
            "fields": fields,
        })
    return merged


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=RIDER_BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def flag_reason(field_name: str, choice: str) -> str | None:
    """Choices that look like import artifacts rather than real vocabulary."""
    if choice.strip().lower() == field_name.strip().lower():
        return "Option name equals the field name — likely a header row imported as data"
    if choice.strip().upper() in {"TRUE", "FALSE"}:
        return "Boolean literal as a select option — likely a bad import/typecast"
    return None


def build(tables: list[dict], out: Path) -> dict:
    wb = Workbook()

    # --- Overview -----------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Table", "Table ID", "Fields", "Select options",
               "Primary field", "Description"])
    for t in tables:
        n_opts = sum(len(f["choices"]) for f in t["fields"])
        primary = next((f["name"] for f in t["fields"]
                        if f["id"] == t["primaryFieldId"]), "")
        ws.append([t["name"], t["id"], len(t["fields"]), n_opts, primary,
                   t["description"]])
    ws.append([])
    ws.append(["Base", "appWHOVZ0QCS0xQ3M"])
    ws.append(["Snapshot taken", date.today().isoformat()])
    ws.append(["Source", "Airtable MCP: list_tables_for_base + get_table_schema"])
    style_header(ws, 6)
    autosize(ws, {1: 30, 2: 22, 3: 9, 4: 14, 5: 20, 6: 80})

    # --- One sheet per table ------------------------------------------------
    for t in tables:
        title = t["name"][:31]  # Excel hard limit
        ws = wb.create_sheet(title)
        ws.append(["#", "Field name", "Field ID", "Type", "Description",
                   "Choices", "# choices", "Config"])
        for f in t["fields"]:
            ws.append([
                f["order"], f["name"], f["id"], f["type"], f["description"],
                " | ".join(f["choices"]),
                len(f["choices"]) or None,
                json.dumps(f["config"]) if f["config"] else "",
            ])
        style_header(ws, 8)
        autosize(ws, {1: 5, 2: 26, 3: 20, 4: 20, 5: 70, 6: 60, 7: 10, 8: 34})

    # --- Select Options (long) ---------------------------------------------
    ws = wb.create_sheet("Select Options")
    ws.append(["Table", "Field", "Field type", "Option", "Option ID", "Colour"])
    for t in tables:
        for f in t["fields"]:
            for name, oid, color in zip(f["choices"], f["choice_ids"],
                                        f["choice_colors"]):
                ws.append([t["name"], f["name"], f["type"], name, oid, color])
    style_header(ws, 6)
    autosize(ws, {1: 28, 2: 24, 3: 16, 4: 34, 5: 24, 6: 16})

    # --- Review Flags -------------------------------------------------------
    ws = wb.create_sheet("Review Flags")
    ws.append(["Table", "Field", "Item", "Flag"])
    flags = 0
    for t in tables:
        for f in t["fields"]:
            for choice in f["choices"]:
                reason = flag_reason(f["name"], choice)
                if reason:
                    ws.append([t["name"], f["name"], choice, reason])
                    flags += 1
            if not f["description"]:
                ws.append([t["name"], f["name"], "(field)",
                           "No description — undocumented in Airtable"])
                flags += 1
    style_header(ws, 4)
    autosize(ws, {1: 28, 2: 26, 3: 34, 4: 74})
    for row in ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.fill = FLAG_FILL

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return {"tables": len(tables),
            "fields": sum(len(t["fields"]) for t in tables),
            "options": sum(len(f["choices"]) for t in tables for f in t["fields"]),
            "flags": flags}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--names", type=Path, required=True,
                    help="raw JSON from list_tables_for_base")
    ap.add_argument("--config", type=Path, required=True,
                    help="raw JSON from get_table_schema")
    ap.add_argument("--out", type=Path, required=True, help="output .xlsx")
    ap.add_argument("--snapshot", type=Path,
                    help="also write the merged schema as JSON here")
    args = ap.parse_args()

    tables = load_merged(args.names, args.config)
    if args.snapshot:
        args.snapshot.parent.mkdir(parents=True, exist_ok=True)
        args.snapshot.write_text(json.dumps(
            {"base": "appWHOVZ0QCS0xQ3M",
             "snapshot_date": date.today().isoformat(),
             "tables": tables}, indent=1) + "\n")

    stats = build(tables, args.out)
    print(f"{args.out} — {stats['tables']} tables, {stats['fields']} fields, "
          f"{stats['options']} select options, {stats['flags']} review flags")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
